"""P4 correction-CLI tests: annotate -> edit -> accept -> draft case, plus
bad-annotation rejection and doc_type dropdown presence.

Runs against the precomputed gold project (data/projects/gold), writing the
draft case into a tmp cases dir so the committed case-11108.json is untouched.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from wbsgen.cases import CaseSource, CaseStatus, load_case
from wbsgen.review import (
    AnnotationError, accept, annotate, read_annotation_xlsx, validate_l1,
    write_annotation_xlsx,
)

GOLD = Path("data/projects/gold")


@pytest.fixture(scope="module", autouse=True)
def _require_gold():
    if not (GOLD / "04_subdoc" / "subdocs.json").exists():
        pytest.skip("gold project not built; run the pipeline first")


# ── annotate produces a workbook with a doc_type dropdown ──────────────

def test_annotate_writes_xlsx_with_dropdown(tmp_path):
    xlsx, yaml = annotate(GOLD, tmp_path / "review")
    assert xlsx.exists() and yaml.exists()

    wb = load_workbook(str(xlsx))
    ws = wb["L1"]
    assert [c.value for c in ws[1]] == ["page_start", "page_end", "doc_type", "title"]
    # at least one data validation covering the doc_type column (C)
    dvs = list(ws.data_validations.dataValidation)
    assert dvs, "no data validation found on annotation sheet"
    assert any("C2" in str(dv.sqref) for dv in dvs)
    assert "CONTRACT_BODY" in dvs[0].formula1


# ── full loop: annotate -> edit one row -> accept -> draft case ────────

def test_annotate_edit_accept_creates_draft_case(tmp_path):
    review_dir = tmp_path / "review"
    xlsx, _ = annotate(GOLD, review_dir)

    # edit one row's title (row 2 = first subdocument)
    wb = load_workbook(str(xlsx))
    ws = wb["L1"]
    ws.cell(row=2, column=4).value = "契約本文(已校正)"
    wb.save(str(xlsx))

    cases_dir = tmp_path / "cases"
    out = accept(GOLD, xlsx, cases_dir, case_id="case-gold-test")
    assert out.exists()

    case = load_case(out)
    assert case.source == CaseSource.CORRECTED
    assert case.status == CaseStatus.DRAFT
    assert case.labels.L1[0].title == "契約本文(已校正)"
    assert len(case.labels.L1) == 10  # gold has 10 subdocuments


# ── bad annotations are rejected without writing a file ───────────────

def test_overlapping_ranges_rejected(tmp_path):
    xlsx = tmp_path / "bad.xlsx"
    write_annotation_xlsx(
        [
            {"page_start": 0, "page_end": 47, "doc_type": "CONTRACT_BODY", "title": "a"},
            {"page_start": 40, "page_end": 63, "doc_type": "ATTACHMENT", "title": "b"},
        ],
        xlsx,
    )
    cases_dir = tmp_path / "cases"
    with pytest.raises(AnnotationError):
        accept(GOLD, xlsx, cases_dir, case_id="case-bad")
    assert not (cases_dir / "case-bad.json").exists()


def test_illegal_doc_type_rejected():
    rows = [{"page_start": 0, "page_end": 5, "doc_type": "NOT_A_TYPE", "title": ""}]
    with pytest.raises(AnnotationError):
        validate_l1(rows)


def test_noncontiguous_gap_rejected():
    rows = [
        {"page_start": 0, "page_end": 5, "doc_type": "CONTRACT_BODY", "title": ""},
        {"page_start": 10, "page_end": 15, "doc_type": "ATTACHMENT", "title": ""},
    ]
    with pytest.raises(AnnotationError):
        validate_l1(rows)


def test_roundtrip_read_matches_written(tmp_path):
    subdocs = [
        {"page_start": 0, "page_end": 5, "doc_type": "CONTRACT_BODY", "title": "本文"},
        {"page_start": 6, "page_end": 9, "doc_type": "ATTACHMENT", "title": "附件"},
    ]
    xlsx = tmp_path / "a.xlsx"
    write_annotation_xlsx(subdocs, xlsx)
    rows = read_annotation_xlsx(xlsx)
    assert len(rows) == 2
    assert rows[0]["doc_type"] == "CONTRACT_BODY"
    assert rows[1]["title"] == "附件"
