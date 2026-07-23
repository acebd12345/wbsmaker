"""Stage 15: Export — xlsx, json, mermaid, csv, coverage.txt."""
from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from ..models import WbsNode


def run(proj_dir: Path, cfg: dict, manifest) -> dict | None:
    wbs_path = proj_dir / "13_merge" / "wbs.json"
    quality_path = proj_dir / "02_quality" / "summary.json"
    subdocs_path = proj_dir / "04_subdoc" / "subdocs.json"
    sections_path = proj_dir / "06_section" / "sections.json"
    tables_path = proj_dir / "07_table" / "tables.json"
    report_path = proj_dir / "14_validate" / "report.json"
    out_dir = proj_dir / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)

    wbs_data = json.loads(wbs_path.read_text(encoding="utf-8"))
    wbs_nodes = [WbsNode(**n) for n in wbs_data]

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    version = "v1"

    # 1. JSON export
    json_path = out_dir / f"wbs_{version}_{timestamp}.json"
    json_path.write_text(
        json.dumps(wbs_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # 2. CSV export
    csv_path = out_dir / f"wbs_{version}_{timestamp}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Code", "Title", "Level", "Type", "Work Items", "Source Pages"])
        for node in wbs_nodes:
            if node.level == 0:
                continue
            indent = "  " * (node.level - 1)
            writer.writerow([
                node.code,
                indent + node.title,
                node.level,
                node.generation_type.value,
                len(node.work_items),
                ",".join(str(p + 1) for p in node.source_pages[:5]),
            ])

    # 3. Mermaid mindmap
    mmd_path = out_dir / f"wbs_{version}_{timestamp}.mmd"
    mmd_lines = ["mindmap", "  root((WBS))"]
    for node in wbs_nodes:
        if node.level == 0:
            continue
        indent = "    " * node.level
        label = node.code + " " + node.title if node.code else node.title
        # Clean label for mermaid (no special chars)
        label = label.replace("(", "").replace(")", "").replace("[", "").replace("]", "")
        mmd_lines.append(f"{indent}{label}")
    mmd_path.write_text("\n".join(mmd_lines), encoding="utf-8")

    # 4. Excel export
    _export_xlsx(out_dir / f"wbs_{version}_{timestamp}.xlsx", wbs_nodes)

    # 5. Coverage report
    coverage_path = out_dir / "coverage.txt"
    _write_coverage(coverage_path, proj_dir, wbs_nodes)

    # 6. Interactive HTML (chart + list views, self-contained)
    _export_html(out_dir / f"wbs_{version}_{timestamp}.html", proj_dir, cfg, wbs_data)

    return None


def _export_html(path: Path, proj_dir: Path, cfg: dict, wbs_data: list[dict]):
    """Inject pipeline data into the visualization template."""
    template_path = Path(__file__).parent.parent / "templates" / "wbs_viz_template.html"
    if not template_path.exists():
        return

    items_path = proj_dir / "10_extract" / "work_items.jsonl"
    sections_path = proj_dir / "06_section" / "sections.json"
    subdocs_path = proj_dir / "04_subdoc" / "subdocs.json"
    quality_path = proj_dir / "02_quality" / "summary.json"
    manifest_path = proj_dir / "manifest.json"

    items: list[dict] = []
    if items_path.exists():
        for line in items_path.read_text(encoding="utf-8").strip().split("\n"):
            if line.strip():
                items.append(json.loads(line))

    sections = json.loads(sections_path.read_text(encoding="utf-8")) if sections_path.exists() else []
    subdocs = json.loads(subdocs_path.read_text(encoding="utf-8")) if subdocs_path.exists() else []
    quality = json.loads(quality_path.read_text(encoding="utf-8")) if quality_path.exists() else {}
    manifest = json.loads(manifest_path.read_text(encoding="utf-8")) if manifest_path.exists() else {}

    sec_title = {s["section_id"]: s["title"] for s in sections}
    subdoc_title = {s["subdoc_id"]: s.get("title", s["subdoc_id"]) for s in subdocs}
    item_by_id = {i["item_id"]: i for i in items}

    def slim(it: dict) -> dict:
        return {
            "id": it["item_id"],
            "desc": it["description"],
            "cat": it["category"],
            "pages": sorted({p + 1 for p in it.get("source_pages", [])}),
            "sec": sec_title.get(it["section_id"], it["section_id"]),
        }

    attached: set[str] = set()
    nodes_out = []
    for n in wbs_data:
        attached.update(n.get("work_items", []))
        nodes_out.append({
            "id": n["node_id"],
            "code": n.get("code", ""),
            "title": n["title"],
            "level": n["level"],
            "parent": n.get("parent_id"),
            "items": [slim(item_by_id[i]) for i in n.get("work_items", []) if i in item_by_id],
        })

    orphans: dict[str, list[dict]] = {}
    for it in items:
        if it["item_id"] in attached:
            continue
        branch = subdoc_title.get(it["subdoc_id"], it["subdoc_id"])
        orphans.setdefault(branch, []).append(slim(it))

    data = {
        "meta": {
            "project": manifest.get("project_id", proj_dir.name),
            "total_pages": quality.get("total_pages"),
            "garbled_pages": quality.get("garbled_text"),
            "image_pages": quality.get("image_only"),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "mode": "mock" if cfg.get("llm", {}).get("mock", True) else "real",
            # relative link from exports/ to the ingested PDF; page numbers
            # become "open PDF at page N" links when it exists
            "pdf_href": (
                "../original/contract.pdf"
                if (proj_dir / "original" / "contract.pdf").exists()
                else None
            ),
        },
        "nodes": nodes_out,
        "orphans": orphans,
        "stats": {
            "total_items": len(items),
            "attached": len(attached & set(item_by_id)),
            "nodes": len(wbs_data),
        },
    }

    html = template_path.read_text(encoding="utf-8")
    # </script> inside JSON strings would terminate the script block early
    payload = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    html = html.replace("__WBS_DATA__", payload, 1)
    path.write_text(html, encoding="utf-8")


def _export_xlsx(path: Path, nodes: list[WbsNode]):
    """Export WBS to Excel with hierarchy indentation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    # Header
    headers = ["Code", "Title", "Level", "Type", "Work Items", "Source Pages"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)

    # Data
    row_num = 2
    for node in nodes:
        if node.level == 0:
            continue

        indent = "  " * (node.level - 1)
        ws.cell(row=row_num, column=1, value=node.code)
        ws.cell(row=row_num, column=2, value=indent + node.title)
        ws.cell(row=row_num, column=3, value=node.level)
        ws.cell(row=row_num, column=4, value=node.generation_type.value)
        ws.cell(row=row_num, column=5, value=len(node.work_items))
        ws.cell(row=row_num, column=6, value=",".join(str(p + 1) for p in node.source_pages[:5]))

        # Bold for level 1
        if node.level == 1:
            for ci in range(1, 7):
                ws.cell(row=row_num, column=ci).font = Font(bold=True)

        row_num += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20

    wb.save(path)


def _write_coverage(path: Path, proj_dir: Path, nodes: list[WbsNode]):
    """Write coverage.txt summary."""
    quality_path = proj_dir / "02_quality" / "summary.json"
    subdocs_path = proj_dir / "04_subdoc" / "subdocs.json"
    sections_path = proj_dir / "06_section" / "sections.json"
    tables_path = proj_dir / "07_table" / "tables.json"

    quality = json.loads(quality_path.read_text(encoding="utf-8")) if quality_path.exists() else {}
    subdocs = json.loads(subdocs_path.read_text(encoding="utf-8")) if subdocs_path.exists() else []
    sections = json.loads(sections_path.read_text(encoding="utf-8")) if sections_path.exists() else []
    tables = json.loads(tables_path.read_text(encoding="utf-8")) if tables_path.exists() else []

    import re
    captioned_tables = [t for t in tables if re.search(r"表\s*\d+", t.get("caption", ""))]

    lines = [
        "=== WBS Coverage Report ===",
        "",
        f"Total pages:    {quality.get('total_pages', '?')}",
        f"Normal text:    {quality.get('normal_text', '?')}",
        f"Image only:     {quality.get('image_only', '?')}",
        f"Garbled text:   {quality.get('garbled_text', '?')}",
        "",
        f"Subdocuments:   {len(subdocs)}",
        f"Sections:       {len(sections)}",
        f"Tables (named): {len(captioned_tables)}",
        f"WBS nodes:      {len([n for n in nodes if n.level > 0])}",
        "",
        "Subdocument breakdown:",
    ]

    for sd in subdocs:
        lines.append(f"  {sd.get('title', sd['subdoc_id']):20s}  p{sd['page_start']+1}-p{sd['page_end']+1}  ({sd['page_count']}p)  [{sd['doc_type']}]")

    path.write_text("\n".join(lines), encoding="utf-8")
