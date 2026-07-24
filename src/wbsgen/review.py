"""Correction CLI core v0 (LEARNING_V0 P4) — L1 subdocument layer only.

Closes the human-in-the-loop correction loop for the subdocument split:
  list     — surface NEEDS_REVIEW / low-confidence signals for a project
  annotate — emit a pre-filled L1 annotation workbook (+ engineer YAML)
  accept   — validate a filled annotation and store it as a *draft* case
             (source=corrected, status=draft)

L2-L4 are reserved in the case schema but intentionally out of scope here.
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .cases import (
    Case, CaseSource, CaseStatus, L1Label, Labels, Provenance,
    compute_fingerprint, dump_case,
)
from .manifest import Manifest
from .models import SubdocType

DOC_TYPES = [t.value for t in SubdocType]
_ANNOTATION_HEADERS = ["page_start", "page_end", "doc_type", "title"]
_BLANK_ROWS = 5


# ── list: surface review signals ───────────────────────────────────────

def collect_review_items(proj_dir: Path) -> dict:
    """Gather NEEDS_REVIEW / low-confidence signals from pipeline artifacts."""
    out: dict = {"needs_review": [], "low_conf_tables": [], "coverage_gaps": []}

    tables_path = proj_dir / "07_table" / "tables.json"
    if tables_path.exists():
        tables = json.loads(tables_path.read_text(encoding="utf-8"))
        for t in tables:
            if t.get("needs_review") or (t.get("merge_confidence", 1.0) < 0.80):
                out["low_conf_tables"].append({
                    "table_id": t.get("table_id"),
                    "merge_confidence": t.get("merge_confidence"),
                    "pages": [t.get("page_start"), t.get("page_end")],
                })

    subdocs_path = proj_dir / "04_subdoc" / "subdocs.json"
    if subdocs_path.exists():
        subdocs = sorted(
            json.loads(subdocs_path.read_text(encoding="utf-8")),
            key=lambda s: s["page_start"],
        )
        prev_end = -1
        for s in subdocs:
            if s["page_start"] > prev_end + 1:
                out["coverage_gaps"].append([prev_end + 1, s["page_start"] - 1])
            prev_end = max(prev_end, s["page_end"])

    report_path = proj_dir / "14_validate" / "report.json"
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))
        for iss in report.get("issues", []):
            if iss.get("severity") in ("NEEDS_REVIEW", "ERROR"):
                out["needs_review"].append(iss)

    return out


# ── annotate: emit pre-filled workbook + YAML ──────────────────────────

def _load_subdocs(proj_dir: Path) -> list[dict]:
    p = proj_dir / "04_subdoc" / "subdocs.json"
    if not p.exists():
        raise FileNotFoundError(f"subdocs not found: {p} (run the pipeline first)")
    return sorted(json.loads(p.read_text(encoding="utf-8")),
                  key=lambda s: s["page_start"])


def write_annotation_xlsx(subdocs: list[dict], out_path: Path) -> None:
    """Write an L1 annotation workbook with a doc_type dropdown."""
    wb = Workbook()
    ws = wb.active
    ws.title = "L1"
    ws.append(_ANNOTATION_HEADERS)
    for s in subdocs:
        ws.append([s["page_start"], s["page_end"], s["doc_type"], s.get("title", "")])
    for _ in range(_BLANK_ROWS):
        ws.append([None, None, None, None])

    # dropdown data validation on the doc_type column (C) for all data rows
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DOC_TYPES) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "doc_type must be one of the listed subdocument types"
    dv.errorTitle = "Invalid doc_type"
    ws.add_data_validation(dv)
    last_row = 1 + len(subdocs) + _BLANK_ROWS
    dv.add(f"C2:C{last_row}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def _yaml_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def write_annotation_yaml(subdocs: list[dict], project: str, out_path: Path) -> None:
    """Emit an engineer-friendly YAML mirror of the annotation (no PyYAML dep)."""
    lines = [f"# L1 subdocument annotation for project: {project}", "subdocs:"]
    for s in subdocs:
        lines.append(f"  - page_start: {s['page_start']}")
        lines.append(f"    page_end: {s['page_end']}")
        lines.append(f"    doc_type: {s['doc_type']}")
        lines.append(f'    title: "{_yaml_escape(s.get("title", ""))}"')
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def annotate(proj_dir: Path, review_dir: Path) -> tuple[Path, Path]:
    subdocs = _load_subdocs(proj_dir)
    xlsx = review_dir / "annotation_L1.xlsx"
    yaml = review_dir / "annotation_L1.yaml"
    write_annotation_xlsx(subdocs, xlsx)
    write_annotation_yaml(subdocs, proj_dir.name, yaml)
    return xlsx, yaml


# ── accept: validate + store as draft case ─────────────────────────────

class AnnotationError(ValueError):
    """Raised when a filled annotation is invalid (no case is written)."""


def read_annotation_xlsx(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(str(xlsx_path))
    ws = wb["L1"] if "L1" in wb.sheetnames else wb.active
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(c is None or str(c).strip() == "" for c in r[:4]):
            continue
        ps, pe, dt, title = (list(r) + [None, None, None, None])[:4]
        rows.append({
            "page_start": ps, "page_end": pe,
            "doc_type": (dt or "").strip() if isinstance(dt, str) else dt,
            "title": (title or "").strip() if isinstance(title, str) else (title or ""),
        })
    return rows


def validate_l1(rows: list[dict]) -> list[L1Label]:
    """Validate legality (doc_type) + contiguity (no gaps / overlaps)."""
    if not rows:
        raise AnnotationError("annotation is empty: at least one subdocument required")

    labels: list[L1Label] = []
    for i, row in enumerate(rows):
        ps, pe, dt = row["page_start"], row["page_end"], row["doc_type"]
        if ps is None or pe is None:
            raise AnnotationError(f"row {i + 1}: page_start/page_end must not be blank")
        try:
            ps, pe = int(ps), int(pe)
        except (TypeError, ValueError):
            raise AnnotationError(f"row {i + 1}: page_start/page_end must be integers")
        if dt not in DOC_TYPES:
            raise AnnotationError(
                f"row {i + 1}: illegal doc_type {dt!r}; must be one of {DOC_TYPES}"
            )
        if ps > pe:
            raise AnnotationError(f"row {i + 1}: page_start {ps} > page_end {pe}")
        labels.append(L1Label(page_start=ps, page_end=pe, doc_type=dt,
                              title=row["title"] or None))

    ordered = sorted(labels, key=lambda l: l.page_start)
    prev_end = None
    for l in ordered:
        if prev_end is not None and l.page_start != prev_end + 1:
            if l.page_start <= prev_end:
                raise AnnotationError(
                    f"page ranges overlap near page {l.page_start} "
                    f"(previous range ended at {prev_end})"
                )
            raise AnnotationError(
                f"page ranges not contiguous: gap between {prev_end} and {l.page_start}"
            )
        prev_end = l.page_end
    return ordered


def accept(proj_dir: Path, xlsx_path: Path, cases_dir: Path,
           case_id: str | None = None) -> Path:
    """Validate a filled annotation and write a draft (corrected) case."""
    m = Manifest(proj_dir)
    if not m.pdf_sha256:
        raise AnnotationError(f"project {proj_dir.name} has no pdf_sha256 in manifest")

    rows = read_annotation_xlsx(xlsx_path)
    labels = validate_l1(rows)  # raises before any file is written

    cid = case_id or f"case-{m.project_id}"
    case = Case(
        case_id=cid,
        pdf_sha256=m.pdf_sha256,
        source=CaseSource.CORRECTED,
        status=CaseStatus.DRAFT,
        answer_schema_version=1,
        labels=Labels(L1=labels),
        fingerprint=compute_fingerprint(proj_dir),
        provenance=Provenance(
            annotator="review-cli",
            annotated_at=date.today().isoformat(),
        ),
    )
    out = cases_dir / f"{cid}.json"
    dump_case(case, out)
    return out
